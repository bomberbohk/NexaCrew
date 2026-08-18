# SPDX-License-Identifier: MIT
"""Enterprise double-entry Accounting (spec §10–§11).

Chart of accounts, balanced journal entries with draft→posted→reversed
lifecycle, financial reports (Trial Balance, P&L, Balance Sheet, GL) and
the accountant-ready tax-season XLSX export."""
from __future__ import annotations

import datetime as dt
import json

# ------------------------------------------------------------
# Default chart of accounts — seeded once per user
# ------------------------------------------------------------
DEFAULT_COA: list[tuple[str, str, str]] = [
    ("1000", "Cash on Hand", "asset"), ("1010", "Bank Checking", "asset"),
    ("1100", "Accounts Receivable", "asset"), ("1200", "Inventory Asset", "asset"),
    ("1500", "Fixed Assets", "asset"), ("1510", "Accumulated Depreciation", "asset"),
    ("2000", "Accounts Payable", "liability"), ("2200", "Sales Tax Payable", "liability"),
    ("2300", "Payroll Liabilities", "liability"), ("2500", "Loans Payable", "liability"),
    ("3000", "Owner's Equity", "equity"), ("3900", "Retained Earnings", "equity"),
    ("4000", "Sales Revenue", "revenue"), ("4100", "Delivery Platform Revenue", "revenue"),
    ("4900", "Other Income", "other income"),
    ("5000", "Cost of Goods Sold", "cogs"), ("5100", "Inventory Shrinkage", "cogs"),
    ("6000", "Rent Expense", "expense"), ("6100", "Utilities", "expense"),
    ("6200", "Payroll Expense", "expense"), ("6300", "Freight In", "expense"),
    ("6400", "Supplies", "expense"), ("6500", "Marketing", "expense"),
    ("6600", "Insurance", "expense"), ("6700", "Repairs & Maintenance", "expense"),
    ("6800", "Processor & Platform Fees", "expense"), ("6900", "Depreciation Expense", "expense"),
    ("7000", "Other Expenses", "other expense"),
]


def seed_coa(db, user_id: str) -> int:
    from .db import Account
    have = {a.number for a in db.query(Account).filter(Account.user_id == user_id).all()}
    n = 0
    for num, name, typ in DEFAULT_COA:
        if num not in have:
            db.add(Account(user_id=user_id, number=num, name=name, type=typ))
            n += 1
    if n:
        db.commit()
    return n


# ------------------------------------------------------------
# Journal entries — always balanced, posted entries immutable
# (corrections via reversal, spec §20/§24)
# ------------------------------------------------------------
def _validate_lines(lines: list[dict]) -> tuple[float, float]:
    if not lines or len(lines) < 2:
        raise ValueError("A journal entry needs at least two lines")
    td = tc = 0.0
    for ln in lines:
        d = float(ln.get("debit") or 0)
        c = float(ln.get("credit") or 0)
        if d < 0 or c < 0:
            raise ValueError("Debits and credits must be non-negative")
        if d and c:
            raise ValueError("A line may have a debit OR a credit, not both")
        if not (ln.get("account") or "").strip():
            raise ValueError("Every line needs an account")
        td += d
        tc += c
    if round(td - tc, 2) != 0:
        raise ValueError(f"Entry is not balanced: debits {td:.2f} ≠ credits {tc:.2f}")
    return round(td, 2), round(tc, 2)


def create_journal(db, user_id: str, lines: list[dict], memo: str = "",
                   source: str = "manual", source_ref: str = "", post: bool = False,
                   at: "dt.datetime | None" = None):
    from .db import JournalEntry
    _validate_lines(lines)
    last = (db.query(JournalEntry).filter(JournalEntry.user_id == user_id)
            .order_by(JournalEntry.number.desc()).first())
    je = JournalEntry(user_id=user_id, number=(last.number if last else 0) + 1,
                      at=at or dt.datetime.utcnow(), memo=memo,
                      lines=json.dumps(lines), source=source, source_ref=source_ref,
                      status="posted" if post else "draft")
    db.add(je)
    db.commit()
    db.refresh(je)
    return je


def reverse_journal(db, user_id: str, je) -> "object":
    if je.status != "posted":
        raise ValueError("Only posted entries can be reversed")
    if je.reversed_by:
        raise ValueError("Entry already reversed")
    lines = json.loads(je.lines or "[]")
    rev = [{"account": ln["account"], "debit": ln.get("credit") or 0,
            "credit": ln.get("debit") or 0, "memo": "REVERSAL: " + (ln.get("memo") or "")}
           for ln in lines]
    r = create_journal(db, user_id, rev, memo=f"Reversal of JE #{je.number}: {je.memo}",
                       source="reversal", source_ref=je.id, post=True)
    je.reversed_by = r.id
    je.status = "reversed"
    db.commit()
    return r


# ------------------------------------------------------------
# Reports
# ------------------------------------------------------------
_NORMAL_DEBIT = {"asset", "cogs", "expense", "other expense"}


def _posted(db, user_id: str, d1=None, d2=None):
    from .db import JournalEntry
    q = db.query(JournalEntry).filter(JournalEntry.user_id == user_id,
                                      JournalEntry.status.in_(["posted", "reversed"]))
    if d1:
        q = q.filter(JournalEntry.at >= d1)
    if d2:
        q = q.filter(JournalEntry.at <= d2)
    return q.order_by(JournalEntry.number).all()


def account_types(db, user_id: str) -> dict[str, str]:
    from .db import Account
    out = {}
    for a in db.query(Account).filter(Account.user_id == user_id).all():
        out[a.number] = a.type
        out[f"{a.number} {a.name}"] = a.type
    return out


def _acct_type(acct: str, types: dict[str, str]) -> str:
    if acct in types:
        return types[acct]
    num = acct.split(" ")[0]
    return types.get(num, "expense")


def trial_balance(db, user_id: str, d1=None, d2=None) -> list[dict]:
    bal: dict[str, list[float]] = {}
    for je in _posted(db, user_id, d1, d2):
        for ln in json.loads(je.lines or "[]"):
            b = bal.setdefault(ln["account"], [0.0, 0.0])
            b[0] += float(ln.get("debit") or 0)
            b[1] += float(ln.get("credit") or 0)
    return [{"account": a, "debit": round(v[0], 2), "credit": round(v[1], 2),
             "balance": round(v[0] - v[1], 2)}
            for a, v in sorted(bal.items())]


def profit_and_loss(db, user_id: str, d1=None, d2=None) -> dict:
    types = account_types(db, user_id)
    rev = cogs = exp = oth_i = oth_e = 0.0
    detail: dict[str, dict[str, float]] = {"revenue": {}, "cogs": {}, "expense": {}}
    for row in trial_balance(db, user_id, d1, d2):
        t = _acct_type(row["account"], types)
        amt = -row["balance"] if t in ("revenue", "other income") else row["balance"]
        if t == "revenue":
            rev += amt
            detail["revenue"][row["account"]] = amt
        elif t == "other income":
            oth_i += amt
        elif t == "cogs":
            cogs += amt
            detail["cogs"][row["account"]] = amt
        elif t in ("expense",):
            exp += amt
            detail["expense"][row["account"]] = amt
        elif t == "other expense":
            oth_e += amt
    gross = rev - cogs
    net = gross - exp + oth_i - oth_e
    return {"revenue": round(rev, 2), "cogs": round(cogs, 2), "gross_profit": round(gross, 2),
            "expenses": round(exp, 2), "other_income": round(oth_i, 2),
            "other_expenses": round(oth_e, 2), "net_income": round(net, 2),
            "detail": {k: {a: round(x, 2) for a, x in v.items()} for k, v in detail.items()}}


def balance_sheet(db, user_id: str, d2=None) -> dict:
    types = account_types(db, user_id)
    assets: dict[str, float] = {}
    liab: dict[str, float] = {}
    eq: dict[str, float] = {}
    for row in trial_balance(db, user_id, None, d2):
        t = _acct_type(row["account"], types)
        if t == "asset":
            assets[row["account"]] = row["balance"]
        elif t == "liability":
            liab[row["account"]] = -row["balance"]
        elif t == "equity":
            eq[row["account"]] = -row["balance"]
    pl = profit_and_loss(db, user_id, None, d2)
    eq["Current Period Net Income"] = pl["net_income"]
    ta = round(sum(assets.values()), 2)
    tl = round(sum(liab.values()), 2)
    te = round(sum(eq.values()), 2)
    return {"assets": {k: round(v, 2) for k, v in assets.items()}, "total_assets": ta,
            "liabilities": {k: round(v, 2) for k, v in liab.items()}, "total_liabilities": tl,
            "equity": {k: round(v, 2) for k, v in eq.items()}, "total_equity": te,
            "balanced": abs(ta - tl - te) < 0.02}


_CASH_WORDS = ("cash", "bank", "checking", "savings", "petty cash", "undeposited")


def is_cash_account(acct: str, types: dict[str, str] | None = None) -> bool:
    """Cash-and-equivalents test used by the Cash Flow Statement."""
    if types is not None and _acct_type(acct, types) != "asset":
        return False
    low = (acct or "").lower()
    return any(w in low for w in _CASH_WORDS)


def classify_cash_activity(counter_accounts: list[str], types: dict[str, str],
                           source: str = "") -> str:
    """Classify a cash movement as operating | investing | financing based on
    the non-cash accounts on the journal entry (indirect classification)."""
    for a in counter_accounts:
        low = a.lower()
        t = _acct_type(a, types)
        if "fixed asset" in low or "equipment" in low or "depreciation" in low:
            return "investing"
        if t == "equity" or "loan" in low or "note payable" in low or "line of credit" in low:
            return "financing"
    return "operating"


def cash_flow(db, user_id: str, d1=None, d2=None) -> dict:
    """Cash Flow Statement (spec §10.3). Walks every posted journal entry,
    takes the net movement on cash/bank accounts and classifies it by the
    counter-accounts. Beginning cash = balance of cash accounts before d1."""
    types = account_types(db, user_id)
    beginning = 0.0
    if d1:
        for row in trial_balance(db, user_id, None, d1 - dt.timedelta(seconds=1)):
            if is_cash_account(row["account"], types):
                beginning += row["balance"]
    buckets = {"operating": 0.0, "investing": 0.0, "financing": 0.0}
    detail: list[dict] = []
    for je in _posted(db, user_id, d1, d2):
        lines = json.loads(je.lines or "[]")
        delta = sum(float(ln.get("debit") or 0) - float(ln.get("credit") or 0)
                    for ln in lines if is_cash_account(ln["account"], types))
        if abs(delta) < 0.005:
            continue
        counter = [ln["account"] for ln in lines if not is_cash_account(ln["account"], types)]
        bucket = classify_cash_activity(counter, types, je.source)
        buckets[bucket] += delta
        detail.append({"journal": je.number, "date": str(je.at or "")[:10],
                       "memo": je.memo, "activity": bucket, "amount": round(delta, 2)})
    net = sum(buckets.values())
    return {"beginning_cash": round(beginning, 2),
            "operating": round(buckets["operating"], 2),
            "investing": round(buckets["investing"], 2),
            "financing": round(buckets["financing"], 2),
            "net_change": round(net, 2),
            "ending_cash": round(beginning + net, 2),
            "detail": detail}


def general_ledger(db, user_id: str, d1=None, d2=None) -> list[dict]:
    rows = []
    running: dict[str, float] = {}
    for je in _posted(db, user_id, d1, d2):
        for ln in json.loads(je.lines or "[]"):
            a = ln["account"]
            running[a] = running.get(a, 0.0) + float(ln.get("debit") or 0) - float(ln.get("credit") or 0)
            rows.append({"date": str(je.at or "")[:10], "journal": je.number,
                         "account": a, "memo": ln.get("memo") or je.memo,
                         "debit": float(ln.get("debit") or 0), "credit": float(ln.get("credit") or 0),
                         "balance": round(running[a], 2), "source": je.source,
                         "ref": je.source_ref or ""})
    return rows


# ------------------------------------------------------------
# Accountant-ready tax-season XLSX export (spec §11)
# ------------------------------------------------------------
def export_xlsx(db, user_id: str, company_name: str, year: int, out_path: str) -> dict:
    """Professional multi-sheet workbook; validates before writing and lists
    problems in Data Exceptions instead of altering records."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    d1 = dt.datetime(year, 1, 1)
    d2 = dt.datetime(year, 12, 31, 23, 59, 59)
    HEAD = Font(bold=True, color="FFFFFF", size=11)
    FILL = PatternFill("solid", fgColor="1F3864")
    TH = Border(bottom=Side(style="thin", color="AAAAAA"))
    MONEY = "#,##0.00"

    wb = Workbook()

    def sheet(title, headers, rows, money_cols=()):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.font = HEAD
            cell.fill = FILL
            cell.border = TH
        for r in rows:
            ws.append(r)
        for ci in money_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, ci).number_format = MONEY
        for ci, h in enumerate(headers, 1):
            width = max(len(str(h)) + 4, *(len(str(ws.cell(r, ci).value or "")) + 2
                                           for r in range(2, min(ws.max_row, 200) + 1))) if ws.max_row > 1 else len(h) + 4
            ws.column_dimensions[get_column_letter(ci)].width = min(width, 60)
        ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        ws.print_title_rows = "1:1"
        return ws

    # 1. Cover
    cover = wb.active
    cover.title = "Cover"
    cover["B2"] = company_name or "Business"
    cover["B2"].font = Font(bold=True, size=20, color="1F3864")
    cover["B3"] = f"Tax-Season Financial Package — Fiscal Year {year}"
    cover["B4"] = f"Generated {dt.datetime.now():%Y-%m-%d %H:%M} · Accrual basis · USD"
    cover["B6"] = "Prepared for review by a CPA / accountant / tax professional."
    cover["B7"] = "This package does not constitute a tax filing."
    cover.column_dimensions["B"].width = 80

    # 2. Executive summary + 3. P&L
    pl = profit_and_loss(db, user_id, d1, d2)
    sheet("Executive Summary", ["Metric", "Amount"],
          [["Revenue", pl["revenue"]], ["Cost of Goods Sold", pl["cogs"]],
           ["Gross Profit", pl["gross_profit"]], ["Operating Expenses", pl["expenses"]],
           ["Other Income", pl["other_income"]], ["Other Expenses", pl["other_expenses"]],
           ["Net Income", pl["net_income"]]], money_cols=(2,))
    pl_rows = [["REVENUE", ""]] + [[a, v] for a, v in pl["detail"]["revenue"].items()] \
        + [["COST OF GOODS SOLD", ""]] + [[a, v] for a, v in pl["detail"]["cogs"].items()] \
        + [["GROSS PROFIT", pl["gross_profit"]], ["EXPENSES", ""]] \
        + [[a, v] for a, v in pl["detail"]["expense"].items()] \
        + [["NET INCOME", pl["net_income"]]]
    sheet("Profit and Loss", ["Account", "Amount"], pl_rows, money_cols=(2,))

    # 4. Balance sheet
    bs = balance_sheet(db, user_id, d2)
    bs_rows = [["ASSETS", ""]] + [[a, v] for a, v in bs["assets"].items()] \
        + [["TOTAL ASSETS", bs["total_assets"]], ["LIABILITIES", ""]] \
        + [[a, v] for a, v in bs["liabilities"].items()] \
        + [["TOTAL LIABILITIES", bs["total_liabilities"]], ["EQUITY", ""]] \
        + [[a, v] for a, v in bs["equity"].items()] \
        + [["TOTAL EQUITY", bs["total_equity"]]]
    sheet("Balance Sheet", ["Account", "Amount"], bs_rows, money_cols=(2,))

    # 5. Trial balance
    tb = trial_balance(db, user_id, d1, d2)
    sheet("Trial Balance", ["Account", "Debit", "Credit", "Balance"],
          [[r["account"], r["debit"], r["credit"], r["balance"]] for r in tb],
          money_cols=(2, 3, 4))

    # 6. General ledger
    gl = general_ledger(db, user_id, d1, d2)
    sheet("General Ledger",
          ["Date", "Journal #", "Account", "Description", "Debit", "Credit",
           "Running Balance", "Source Module", "Audit Reference"],
          [[r["date"], r["journal"], r["account"], r["memo"], r["debit"], r["credit"],
            r["balance"], r["source"], r["ref"]] for r in gl],
          money_cols=(5, 6, 7))

    # 7. Data exceptions — validation results, never mutates records
    exceptions: list[list] = []
    for je_row in tb:
        pass
    from .db import JournalEntry, VendorInvoice
    drafts = (db.query(JournalEntry)
              .filter(JournalEntry.user_id == user_id, JournalEntry.status == "draft").count())
    if drafts:
        exceptions.append(["Draft journal entries", f"{drafts} unposted draft entr(ies) excluded from reports"])
    unconfirmed = (db.query(VendorInvoice)
                   .filter(VendorInvoice.user_id == user_id,
                           VendorInvoice.status.notin_(["posted", "rejected"])).count())
    if unconfirmed:
        exceptions.append(["Unverified vendor invoices", f"{unconfirmed} invoice(s) pending human verification — not included in AP"])
    if not bs["balanced"]:
        exceptions.append(["Balance sheet", "Assets ≠ Liabilities + Equity — investigate before filing"])
    if not exceptions:
        exceptions.append(["None", "No validation exceptions detected"])
    sheet("Data Exceptions", ["Issue", "Detail"], exceptions)

    sheet("Export Notes", ["Note"], [
        ["Basis: Accrual. Currency: USD. All amounts from posted journal entries only."],
        ["Corrections to posted entries are made via reversal entries, never edits."],
        ["Vendor invoices post to AP only after mandatory field-by-field human verification."],
        ["Generated by NexaCrew — checksum of this file is recorded in the audit log."]])

    wb.save(out_path)
    import hashlib
    checksum = hashlib.sha256(open(out_path, "rb").read()).hexdigest()
    return {"path": out_path, "sheets": len(wb.sheetnames), "gl_rows": len(gl),
            "checksum": checksum, "exceptions": len(exceptions)}
